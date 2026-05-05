from __future__ import annotations

import csv
import io
import re
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from pypdf import PdfReader


class _HTMLTextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_title = False
        self.skip_depth = 0
        self.title_parts: List[str] = []
        self.text_parts: List[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "title":
            self.in_title = True
        if tag in {"script", "style", "noscript"}:
            self.skip_depth += 1

    def handle_endtag(self, tag):
        if tag == "title":
            self.in_title = False
        if tag in {"script", "style", "noscript"} and self.skip_depth:
            self.skip_depth -= 1

    def handle_data(self, data):
        text = " ".join(data.split())
        if not text or self.skip_depth:
            return
        if self.in_title:
            self.title_parts.append(text)
        else:
            self.text_parts.append(text)


SECTOR_KEYWORDS = [
    ("AI", ["artificial intelligence", "machine learning", "foundation model", "agentic", "ai copilot"]),
    ("Cybersecurity", ["cybersecurity", "threat detection", "appsec", "security posture", "vulnerability"]),
    ("Healthtech", ["healthcare", "clinical", "patient", "medtech", "digital health"]),
    ("Climate Infra", ["climate", "methane", "decarbon", "emissions", "energy transition"]),
    ("Compliance", ["compliance", "governance", "risk management", "grc", "audit"]),
    ("Fintech", ["payments", "banking", "lending", "insurance", "fintech"]),
    ("Developer Tools", ["developer platform", "devops", "api platform", "software engineering"]),
]

SUBSECTOR_KEYWORDS = [
    ("B2B SaaS", ["subscription software", "saas", "b2b software"]),
    ("Developer security", ["developer security", "code security", "application security"]),
    ("IP workflow", ["intellectual property", "patent workflow", "patent drafting"]),
    ("Industrial sensing", ["industrial sensing", "sensor platform", "industrial monitoring"]),
    ("Human interface", ["human interface", "voice interface", "human computer interaction"]),
    ("GRC", ["grc", "compliance automation", "governance risk compliance"]),
]

BUSINESS_MODEL_KEYWORDS = [
    ("B2B SaaS", ["saas", "subscription", "annual recurring revenue", "arr"]),
    ("Marketplace", ["marketplace", "buyers and sellers"]),
    ("Platform", ["platform", "ecosystem"]),
    ("Hardware + software", ["hardware and software", "device + software", "sensor platform"]),
    ("Services", ["consulting", "professional services", "agency"]),
]

COUNTRY_CITY_HINTS = [
    ("Belgium", ["ghent", "brussels", "antwerp", "belgium"]),
    ("Netherlands", ["amsterdam", "rotterdam", "utrecht", "netherlands"]),
    ("United Kingdom", ["london", "manchester", "united kingdom", "uk", "england"]),
    ("Germany", ["berlin", "munich", "hamburg", "germany"]),
    ("France", ["paris", "lyon", "france"]),
    ("Switzerland", ["zurich", "geneva", "switzerland"]),
    ("Denmark", ["copenhagen", "denmark"]),
]

EMAIL_RE = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.I)
ROUND_RE = re.compile(r"\b(pre-seed|seed|series\s+[a-z]|series\s+[a-z]\+?)\b", re.I)
OWNERSHIP_RE = re.compile(r"(\d{1,2}(?:\.\d+)?)\s?%\s+(?:ownership|stake)", re.I)
PERCENT_FIELD_PATTERNS = {
    "revenue_growth_pct": [
        r"(?:revenue growth|arr growth|growth rate|yoy growth|year[- ]over[- ]year growth)[^\d]{0,20}(\d{1,3}(?:\.\d+)?)\s?%",
    ],
    "gross_margin_pct": [
        r"(?:gross margin)[^\d]{0,20}(\d{1,3}(?:\.\d+)?)\s?%",
    ],
    "ebitda_margin_pct": [
        r"(?:ebitda margin)[^\d]{0,20}(\d{1,3}(?:\.\d+)?)\s?%",
    ],
    "rule_of_40_pct": [
        r"(?:rule of 40)[^\d]{0,20}(\d{1,3}(?:\.\d+)?)\s?%",
    ],
}


def _normalize_whitespace(value: str) -> str:
    return " ".join((value or "").split()).strip()


def _truncate(value: str, limit: int) -> str:
    value = _normalize_whitespace(value)
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def _read_website(url: str) -> Dict[str, str]:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0 VCCCRM Autofill"})
    with urlopen(request, timeout=12) as response:
        raw = response.read()
        encoding = response.headers.get_content_charset() or "utf-8"
    html = raw.decode(encoding, errors="replace")
    parser = _HTMLTextParser()
    parser.feed(html)
    meta = ""
    match = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']+)["\']', html, re.I)
    if match:
        meta = _normalize_whitespace(match.group(1))
    return {
        "title": _normalize_whitespace(" ".join(parser.title_parts)),
        "meta_description": meta,
        "text": _truncate(" ".join(parser.text_parts), 10000),
    }


def _read_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    text_parts = []
    for page in reader.pages[:12]:
        text_parts.append(page.extract_text() or "")
    return _truncate("\n".join(text_parts), 15000)


def _read_spreadsheet(path: Path) -> str:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: List[str] = []
    for sheet_name in workbook.sheetnames[:4]:
        sheet = workbook[sheet_name]
        lines.append(f"[Sheet] {sheet_name}")
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
            values = [_normalize_whitespace(str(cell)) for cell in row if cell not in (None, "")]
            if not values:
                continue
            lines.append(" | ".join(values[:10]))
            row_count += 1
            if row_count >= 12:
                break
    return _truncate("\n".join(lines), 15000)


def _read_csv_file(path: Path) -> str:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        lines = []
        for index, row in enumerate(reader):
            if index >= 20:
                break
            values = [_normalize_whitespace(cell) for cell in row if _normalize_whitespace(cell)]
            if values:
                lines.append(" | ".join(values[:10]))
    return _truncate("\n".join(lines), 12000)


def _shared_document_path(storage_path: str, shared_documents_dir: Path) -> Optional[Path]:
    if not storage_path:
        return None
    if storage_path.startswith("/shared-documents/"):
        return shared_documents_dir / Path(storage_path).name
    candidate = Path(storage_path)
    if candidate.exists():
        return candidate
    return None


def _collect_document_texts(documents: Iterable[dict], shared_documents_dir: Path) -> Dict[str, str]:
    deck_texts: List[str] = []
    financial_texts: List[str] = []
    warnings: List[str] = []
    for document in documents:
        local_path = _shared_document_path(document.get("storage_path") or "", shared_documents_dir)
        if not local_path or not local_path.exists():
            continue
        category = (document.get("document_category") or "other").lower()
        suffix = local_path.suffix.lower()
        text = ""
        try:
            if suffix == ".pdf":
                text = _read_pdf(local_path)
            elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                text = _read_spreadsheet(local_path)
            elif suffix == ".csv":
                text = _read_csv_file(local_path)
            elif suffix in {".txt", ".md"}:
                text = _truncate(local_path.read_text(encoding="utf-8", errors="replace"), 12000)
        except Exception as error:  # pragma: no cover - best-effort extraction
            warnings.append(f"Could not parse {local_path.name}: {error}")
        if not text:
            continue
        if category == "deck":
            deck_texts.append(text)
        elif category == "financials":
            financial_texts.append(text)
    return {
        "deck_text": _truncate("\n\n".join(deck_texts), 18000),
        "financial_text": _truncate("\n\n".join(financial_texts), 18000),
        "warnings": warnings,
    }


def _pick_keyword_label(text: str, options: List[tuple[str, List[str]]]) -> str:
    lowered = text.lower()
    for label, keywords in options:
        if any(keyword in lowered for keyword in keywords):
            return label
    return ""


def _detect_country_and_city(text: str) -> Dict[str, str]:
    lowered = text.lower()
    for country, hints in COUNTRY_CITY_HINTS:
        for hint in hints:
            if hint in lowered:
                city = hint.title() if hint not in {"belgium", "netherlands", "united kingdom", "uk", "england", "germany", "france", "switzerland", "denmark"} else ""
                return {"hq_country": country, "hq_city": city}
    return {"hq_country": "", "hq_city": ""}


def _extract_money_candidates(text: str) -> List[float]:
    matches = re.findall(r"(?:€|\$|eur\s+)?(\d[\d.,]*)(?:\s?(m|mn|million|b|bn|billion|k|thousand))?", text, re.I)
    values = []
    for raw_number, scale in matches:
        try:
            number = float(raw_number.replace(",", ""))
        except ValueError:
            continue
        scale_key = (scale or "").lower()
        if scale_key in {"k", "thousand"}:
            number *= 1_000
        elif scale_key in {"m", "mn", "million"}:
            number *= 1_000_000
        elif scale_key in {"b", "bn", "billion"}:
            number *= 1_000_000_000
        values.append(number)
    return values


def _extract_metric_money(text: str, labels: List[str]) -> str:
    for label in labels:
        pattern = rf"{label}[^\d]{{0,25}}(?:€|\$|eur\s+)?(\d[\d.,]*)(?:\s?(m|mn|million|b|bn|billion|k|thousand))?"
        match = re.search(pattern, text, re.I)
        if not match:
            continue
        raw_number, scale = match.groups()
        try:
            number = float(raw_number.replace(",", ""))
        except ValueError:
            continue
        scale_key = (scale or "").lower()
        if scale_key in {"k", "thousand"}:
            number *= 1_000
        elif scale_key in {"m", "mn", "million"}:
            number *= 1_000_000
        elif scale_key in {"b", "bn", "billion"}:
            number *= 1_000_000_000
        return str(int(number))
    return ""


def _extract_percent_metric(text: str, field_name: str) -> str:
    for pattern in PERCENT_FIELD_PATTERNS.get(field_name, []):
        match = re.search(pattern, text, re.I)
        if match:
            return match.group(1)
    return ""


def _extract_financial_metrics(text: str) -> Dict[str, str]:
    metrics = {
        "annual_recurring_revenue": _extract_metric_money(text, ["arr", "annual recurring revenue", "revenue run[- ]rate", "revenue"]),
        "monthly_burn": _extract_metric_money(text, ["monthly burn", "burn rate", "cash burn"]),
        "revenue_growth_pct": _extract_percent_metric(text, "revenue_growth_pct"),
        "gross_margin_pct": _extract_percent_metric(text, "gross_margin_pct"),
        "ebitda_margin_pct": _extract_percent_metric(text, "ebitda_margin_pct"),
        "rule_of_40_pct": _extract_percent_metric(text, "rule_of_40_pct"),
        "cash_runway_months": "",
        "financials_updated_at": "",
    }
    runway_match = re.search(r"(\d{1,2}(?:\.\d+)?)\s+(?:months?|mos?)\s+(?:of\s+)?runway", text, re.I)
    if runway_match:
        metrics["cash_runway_months"] = runway_match.group(1)
    date_match = re.search(r"\b(as of|updated|financials dated)\s+([A-Z][a-z]{2,9}\s+\d{4})\b", text)
    if date_match:
        metrics["financials_updated_at"] = date_match.group(2)
    return metrics


def _extract_round(text: str) -> str:
    match = ROUND_RE.search(text)
    if not match:
        return ""
    value = match.group(1).replace("series", "Series").replace("pre-seed", "Pre-Seed").replace("seed", "Seed")
    return value.title().replace("Series ", "Series ")


def _extract_contact(text: str) -> Dict[str, str]:
    emails = EMAIL_RE.findall(text)
    email = emails[0] if emails else ""
    name = ""
    title = ""
    name_match = re.search(r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)\s*[-,|]\s*(CEO|CFO|COO|Founder|Managing Director|VP Finance|Chief Executive Officer)\b", text)
    if name_match:
        name = name_match.group(1)
        title = name_match.group(2)
    return {
        "primary_contact_name": name,
        "primary_contact_title": title,
        "primary_contact_email": email,
        "primary_contact_phone": "",
    }


def _compose_investment_thesis(website: Dict[str, str], deck_text: str, financial_text: str) -> str:
    parts = []
    if website.get("meta_description"):
        parts.append(website["meta_description"])
    elif website.get("title"):
        parts.append(f"Company narrative from website: {website['title']}.")
    sector = _pick_keyword_label(" ".join([website.get("text", ""), deck_text, financial_text]), SECTOR_KEYWORDS)
    business_model = _pick_keyword_label(" ".join([website.get("text", ""), deck_text]), BUSINESS_MODEL_KEYWORDS)
    round_name = _extract_round("\n".join([deck_text, financial_text]))
    if sector or business_model:
        parts.append(f"Likely profile: {sector or 'Sector unclear'}{f' / {business_model}' if business_model else ''}.")
    if round_name:
        parts.append(f"Current financing context appears to reference {round_name}.")
    return _truncate(" ".join(parts), 1200)


def _compose_key_concerns(deck_text: str, financial_text: str, has_financials: bool, contact: Dict[str, str]) -> str:
    concerns = []
    if not has_financials:
        concerns.append("Financial package is missing or could not be parsed.")
    if not contact.get("primary_contact_email"):
        concerns.append("Primary contact email was not confidently detected from available sources.")
    if not _extract_round("\n".join([deck_text, financial_text])):
        concerns.append("Round context is still unclear from the available materials.")
    money_values = _extract_money_candidates("\n".join([deck_text, financial_text]))
    if not money_values:
        concerns.append("Raise size and valuation were not clearly stated in the current sources.")
    return _truncate(" ".join(concerns), 1200)


def run_pipeline_record_autofill(
    *,
    record: dict,
    organization: dict,
    documents: List[dict],
    shared_documents_dir: Path,
) -> dict:
    website_data = {"title": "", "meta_description": "", "text": ""}
    warnings: List[str] = []
    website_url = organization.get("website") or ""
    if website_url:
        try:
            website_data = _read_website(website_url)
        except URLError as error:
            warnings.append(f"Could not fetch website: {error}")
        except Exception as error:  # pragma: no cover - best-effort extraction
            warnings.append(f"Could not parse website: {error}")

    document_texts = _collect_document_texts(documents, shared_documents_dir)
    warnings.extend(document_texts.get("warnings", []))
    deck_text = document_texts.get("deck_text", "")
    financial_text = document_texts.get("financial_text", "")
    combined_text = "\n".join(filter(None, [website_data.get("text"), deck_text, financial_text]))

    sector = _pick_keyword_label(combined_text, SECTOR_KEYWORDS)
    subsector = _pick_keyword_label(combined_text, SUBSECTOR_KEYWORDS)
    business_model = _pick_keyword_label(combined_text, BUSINESS_MODEL_KEYWORDS)
    location = _detect_country_and_city(combined_text)
    contact = _extract_contact("\n".join(filter(None, [website_data.get("text"), deck_text])))
    round_name = _extract_round("\n".join([deck_text, financial_text]))
    money_values = sorted(_extract_money_candidates("\n".join([deck_text, financial_text])))
    ownership_match = OWNERSHIP_RE.search("\n".join([deck_text, financial_text]))

    valuation_min = ""
    valuation_max = ""
    ticket_target = ""
    if money_values:
        if len(money_values) >= 2:
            valuation_min = str(int(money_values[-2]))
            valuation_max = str(int(money_values[-1]))
        else:
            valuation_max = str(int(money_values[-1]))
            ticket_target = str(int(money_values[0]))

    if not ticket_target and money_values:
        smaller_values = [value for value in money_values if value <= 50_000_000]
        if smaller_values:
            ticket_target = str(int(smaller_values[0]))

    financial_metrics = _extract_financial_metrics("\n".join([deck_text, financial_text]))
    concerns = _compose_key_concerns(deck_text, financial_text, bool(financial_text), contact)
    thesis = _compose_investment_thesis(website_data, deck_text, financial_text)

    risk_flags = []
    if not financial_text:
        risk_flags.append("missing_financials")
    if not contact.get("primary_contact_email"):
        risk_flags.append("missing_contact_details")

    organization_updates = {
        "name": organization.get("name") or record.get("organization_name") or "",
        "website": website_url,
        "description": website_data.get("meta_description") or _truncate(website_data.get("text", ""), 1200),
        "sector_primary": sector,
        "subsector": subsector,
        "business_model": business_model,
        "geography": location.get("hq_country") or organization.get("geography") or "",
        "hq_city": location.get("hq_city"),
        "hq_country": location.get("hq_country"),
    }
    record_updates = {
        "round_name": round_name,
        "ticket_size_target": ticket_target,
        "ownership_target_pct": ownership_match.group(1) if ownership_match else "",
        "valuation_min": valuation_min,
        "valuation_max": valuation_max,
        "annual_recurring_revenue": financial_metrics["annual_recurring_revenue"],
        "revenue_growth_pct": financial_metrics["revenue_growth_pct"],
        "gross_margin_pct": financial_metrics["gross_margin_pct"],
        "ebitda_margin_pct": financial_metrics["ebitda_margin_pct"],
        "rule_of_40_pct": financial_metrics["rule_of_40_pct"],
        "monthly_burn": financial_metrics["monthly_burn"],
        "cash_runway_months": financial_metrics["cash_runway_months"],
        "financials_updated_at": financial_metrics["financials_updated_at"],
        "investment_thesis": thesis,
        "key_concerns": concerns,
        "risk_flags": risk_flags,
    }
    source_summary = {
        "website_used": bool(website_url and website_data.get("text")),
        "deck_documents": [item.get("file_name") for item in documents if (item.get("document_category") or "").lower() == "deck"],
        "financial_documents": [item.get("file_name") for item in documents if (item.get("document_category") or "").lower() == "financials"],
        "warnings": warnings,
    }
    return {
        "organization": organization_updates,
        "contact": contact,
        "pipeline_record": record_updates,
        "source_summary": source_summary,
    }
